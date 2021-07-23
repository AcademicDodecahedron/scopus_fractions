import openpyxl
import json
import re
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from argparse import ArgumentParser
from lib import tabular, blockcsv
from lib.model import RequestResult

def parse_args():
    argp = ArgumentParser()
    argp.add_argument('template', help='template .xlsx file')
    argp.add_argument('id', help='university id (AF-ID)')
    argp.add_argument('--fractions')
    argp.add_argument('--publications')
    argp.add_argument('-o', '--out', default='out.xlsx', help='output .xlsx file')
    return argp.parse_args()
args = parse_args()

wb = openpyxl.load_workbook(args.template)
id_aff_trans = wb['id_aff_trans']
fractions = wb['фракции']
publications = wb['публикации']

id_aff_trans.append([args.id, args.id])

if args.fractions is not None:
    print('Reading', args.fractions)
    with open(args.fractions) as file:
        for row_index, row in enumerate(tabular.Reader(file, RequestResult)):
            if not row.ok:
                print(f"Skipping row {row_index}: Failed Request")
                continue
            entries = json.loads(row.response)['search-results'].get('entry', [])
            for entry in entries:
                eid = entry['eid']
                author_list = entry.get('author', [])
                count_au = len(author_list)
                for author in author_list:
                    au_id = author['authid']
                    aff_list = author.get('afid', [])
                    count_aff = len(aff_list)
                    for aff in aff_list:
                        aff_id = aff['$']
                        fractions.append([
                            eid,
                            au_id,
                            aff_id,
                            count_aff,
                            count_au,
                            '=IFERROR(VLOOKUP(фракции[aff],id_aff_trans[],2,0),"")',
                            '=1/фракции[count_au]/фракции[count_aff]'
                        ])

INT_RG = re.compile(r'\d+')
FLOAT_RG = re.compile(r'\d+\.\d+')
def try_convert_numeric(value: str):
    if INT_RG.fullmatch(value) is not None:
        return int(value)
    elif FLOAT_RG.fullmatch(value) is not None:
        return float(value)
    return value

PUBLICATION_FORMULAS = [
    '=IFERROR(VLOOKUP(публикации[[#This Row],[EID]],фракцииУниверситет[],2,0),"")',
    '=IFERROR(VLOOKUP(публикации[[#This Row],[Year]],snipThresholds[],2,0),0)',
    '=IF(публикации[[#This Row],[SNIP (publication year)]]>=публикации[[#This Row],[snipThresholds]],"Q1 Q2","Q3 Q4 n/a")'
]

if args.publications is not None:
    print('Reading ', args.publications)
    with open(args.publications) as file:
        reader = blockcsv.reader(file, '"Title","Authors","Number of Authors"')

        headings = next(reader) + ['фракции', 'snipThresholds', 'критерийотбора']
        publications.append(headings)
        for row in reader:
            row_converted = list(map(try_convert_numeric, row))
            publications.append(row_converted + PUBLICATION_FORMULAS)

def get_cell_range(sheet):
    return "A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
def make_table(sheet, name):
    table = Table(displayName=name, ref=get_cell_range(sheet))
    sheet.add_table(table)
def update_table(sheet, name):
    table = sheet.tables[name]
    table.ref = get_cell_range(sheet)

update_table(id_aff_trans, 'id_aff_trans')
update_table(fractions, 'фракции')
make_table(publications, 'публикации')
wb.save(args.out)
