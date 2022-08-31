from argparse import ArgumentParser
import json
import re
from typing import TextIO
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from safetywrap import Option, Result

from lib import tabular
from lib.model import RequestResult


def update_or_create_table(sheet, name):
    table = sheet.tables.get(name, None)
    if not table:
        table = Table(displayName=name)
        sheet.add_table(table)

    cell_range = "A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
    table.ref = cell_range


@dataclass
class WorkbookSheets:
    wb: Workbook
    record: Worksheet
    affiliation: Worksheet
    author: Worksheet
    fraction: Worksheet
    id_aff_trans: Worksheet

    @staticmethod
    def new():
        wb = Workbook()

        record: Worksheet = wb.active
        record.title = "record"
        record.append(
            [
                "eid",
                "Year",
                "DOI",
                "citedby-count",
                "TypeSource",
                "source-id",
                "TypeRecordShort",
                "TypeRecord",
                "year_extract",
            ]
        )

        affiliation: Worksheet = wb.create_sheet("affiliation")
        affiliation.append(
            ["eid", "afid", "id_aff_trans", "OrgName", "City", "Country"]
        )

        author: Worksheet = wb.create_sheet("author")
        author.append(["eid", "authid", "AuthorName", "given-name", "initials"])

        fraction: Worksheet = wb.create_sheet("fraction")
        fraction.append(
            ["eid", "au_id", "aff", "count_aff", "count_au", "id_aff_trans", "fraction"]
        )

        id_aff_trans: Worksheet = wb.create_sheet("id_aff_trans")
        id_aff_trans.append(["id_aff структуры", "id_aff расчитываемой организации"])

        return WorkbookSheets(wb, record, affiliation, author, fraction, id_aff_trans)

    def save(self, filename: str):
        update_or_create_table(self.record, "record")
        update_or_create_table(self.affiliation, "affiliation")
        update_or_create_table(self.author, "author")
        update_or_create_table(self.fraction, "fraction")
        update_or_create_table(self.id_aff_trans, "id_aff_trans")
        self.wb.save(filename)


YEAR_REGEX = re.compile(r"\d\d\d\d")


def fill_workbook(fractions: TextIO, uni_id: str):
    wb = WorkbookSheets.new()
    wb.id_aff_trans.append([uni_id, uni_id])

    for row_index, row in enumerate(tabular.Reader(fractions, RequestResult)):
        if not row.ok:
            print(f"Skipping row {row_index}: Failed Request")
            continue

        entries = json.loads(row.response)["search-results"].get("entry", [])
        for entry in entries:
            eid = entry["eid"]
            year = entry.get("prism:coverDisplayDate", None)
            wb.record.append(
                [
                    eid,
                    year,
                    entry.get("prism:doi", None),
                    entry.get("citedby-count", None),
                    entry.get("prism:aggregationType", None),
                    entry.get("source-id", None),
                    entry.get("subtype", None),
                    entry.get("subtypeDescription", None),
                    Option.of(year)
                    .flatmap(lambda year: Option.of(YEAR_REGEX.search(year)))
                    .map(lambda match: match.group(0))
                    .flatmap(lambda text: Result.of(int, text).ok())
                    .unwrap_or(None),
                ]
            )
            for affiliation in entry.get("affiliation", []):
                wb.affiliation.append(
                    [
                        eid,
                        affiliation["afid"],
                        '=IFERROR(VLOOKUP(affiliation[afid],id_aff_trans,2,0),"")',
                        affiliation.get("affilname", None),
                        affiliation.get("affiliation-city", None),
                        affiliation.get("affiliation-country", None),
                    ]
                )

            author_list = entry.get("author", [])
            count_authors = len(author_list)
            for author in author_list:
                authid = author["authid"]
                wb.author.append(
                    [
                        eid,
                        authid,
                        author.get("authname", None),
                        author.get("given-name", None),
                        author.get("initials", None),
                    ]
                )

                afid_list = author.get("afid", [])
                count_aff = len(afid_list)
                for afid in afid_list:
                    wb.fraction.append(
                        [
                            eid,
                            authid,
                            afid["$"],
                            count_aff,
                            count_authors,
                            '=IFERROR(VLOOKUP(fraction[aff],id_aff_trans[],2,0),"")',
                            "=1/fraction[count_au]/fraction[count_aff]",
                        ]
                    )

    return wb


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("fractions", help="text file generated by download.py")
    parser.add_argument("id", help="university id (AF-ID)")
    parser.add_argument("-o", "--out", default="out.xlsx", help="output .xlsx file")
    args = parser.parse_args()

    with open(args.fractions, encoding="utf-8") as fractions:
        wb = fill_workbook(fractions, args.id)

    wb.save(args.out)
